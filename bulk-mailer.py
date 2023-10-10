#!/usr/bin/env python3
"""Bulk emailer for EngineeringX course stakeholders/customers
Originally from https://levelup.gitconnected.com/sending-bulk-emails-via-python-4592b7ee57a5
For sending while not on the RU network, you will need to use the SSL/TLS capability.  DON'T SHARE YOUR PASSWORD!
The addresses will be loaded from an .xlsx file
Alternative: Python mailmerge https://pypi.org/project/mailmerge

Ubuntu install
  sudo apt install python3-openpyxl

WARNING:  This program is a work in progress and may break at any time.
TODO:  move credentials to another file
TODO:  command line options and interface for sending
"""

import os
import os.path
# import sys
# import re
import argparse
#import subprocess
import sys
#import pprint
import logging
import configparser #ConfigParser in py2 configparser in py3
import re

# Python Excel library
# Doc: https://openpyxl.readthedocs.io/en/stable/tutorial.html 
from openpyxl import load_workbook

# Import smtplib for our actual email sending function
import smtplib
 
# Helper email modules 
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders


if __name__ == '__main__':
    # http://stackoverflow.com/questions/8299270/ultimate-answer-to-relative-python-imports
    # relative imports do not work when we run this module directly
    PACK_DIR = os.path.dirname(os.path.join(os.getcwd(), __file__))
    ADDTOPATH = os.path.normpath(os.path.join(PACK_DIR, '..'))
    # add more .. depending upon levels deep
    sys.path.append(ADDTOPATH)

SCRIPTPATH = os.path.dirname(os.path.realpath(__file__))
DEVNULL = open(os.devnull, 'wb')

class BulkMailer(object):
    """Objectify the bulk mailer functionality"""

    def __init__(self, args):
        """Setup configuration"""
        self.args = args
        # setup logger
        logpath = 'mycanvas.log'
        floglevel = logging.DEBUG
        cloglevel = logging.INFO
        self.log.setLevel(floglevel)
        self.log.addHandler(logging.FileHandler(logpath))
        console_handler = logging.StreamHandler(sys.stderr)
        console_handler.setLevel(cloglevel)
        self.log.addHandler(console_handler)
        self.log.info("Logging to %s", logpath)
        self.log.debug("File logging set at %s", floglevel)
        self.log.debug("Console logging level at %s", cloglevel)

        # setup the config file
        configpaths = [args.configfile, "./teaching-tools.ini", "~/.teaching-tools.ini"]
        cp = configparser.SafeConfigParser()
        configfile = None
        def checkfile(filepath):
            return os.path.isfile(filepath) and os.access(filepath, os.R_OK)
        for configpath in configpaths:
            self.log.debug("Looking for config file in %s", configpath)
            if configpath is None:
                continue
            configpath = os.path.expanduser(configpath)  #deal with ~
            if not checkfile(configpath):
                continue
            configfile = cp.read([configpath])
            self.log.info("Using configuration file at %s", configpath)
            break

        if configfile is None:
            self.log.error("Error:  Couldn't find a configuration file.")
            sys.exit()

        def config2dict(section):
            """convert a parsed configuration into a dict for storage"""
            config = {setarg: setval for setarg, setval
                  in cp.items(section)}
            for setarg, setval in config.items():
                self.log.debug('setting: %s = %s', setarg, setval)
            return config
        config = {}
        for section in ['api', 'courses']:
            config[section] = config2dict(section)
        self.config = config
        self.configfile = configfile

    def load_senderdb(self, sheet=None):
        """Excel file with sender and customization info"""

        self.wb = load_workbook(filename=filepath)
        # which sheet?  Assume first if not specified
        ws = self.wb.active # TODO need to change to sheet 0        
        if sheet:
            ws = self.wb[sheet]
        firstcol = []
        
        for row in range(1, lastrow):
            val = ws.cell(column=1, row=row).value
            if val is None:
                next
            else:
                firstcol.append(val)
        self.logger.debug(firstcol)
        # convert for orderdict for later
        codes = OrderedDict()
        for k in firstcol:
            codes[k] = startval
        self.codes = codes

        
    def prepare_email(self):
        """Setup and check the email body"""
        # sender email address
        email_user = 'foley@ru.is'
        
        # sender email passowrd for login purposes
        email_password = 'SENDER_EMAIL_PASSWORD'
        
        # list of users to whom email is to be sent
        email_send = ['foley@ru.is']
        subject = 'Test Subject'
        msg = MIMEMultipart()
        msg['From'] = email_user
        # converting list of recipients into comma separated string
        msg['To'] = ", ".join(email_send)
        msg['Subject'] = subject
        body = 'THIS IS A TEST'
        msg.attach(MIMEText(body,'plain'))
        text = msg.as_string()

    def send_emails(self):
        #server = smtplib.SMTP('smtp.ru.is',587)
        server = smtplib.SMTP('smtp.ru.is')
        
        ## TLS/SSL support (if needed)
        #server.starttls()
        #server.login(email_user,email_password)
        
        server.sendmail(email_user,email_send,text)
        server.quit()

if __name__ == '__main__':
    PARSER = argparse.ArgumentParser(
        description='Excel-db Bulk Mailer for EngineeringX')
    PARSER.add_argument('--version', action="version", version="%(prog)s 0.1")  #version init was depricated
    PARSER.add_argument('emailtemplate',#required!
                        help='email template')
    PARSER.add_argument('-c', '--configfile',
                        help='configuration file location (override)')
    parser.add_argument('--log', default="INFO",
                        help='Console log level:  Number or DEBUG, INFO, WARNING, ERROR')
    parser.add_argument('--test',
                        help='Test mode:  No email sent, just messages')
    
    ARGS=PARSER.parse_args()

    BM = BulkMailer(args=ARGS)
    BM.load_senderdb()
