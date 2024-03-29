#!/usr/bin/python3
## Abstraction for working with Rubrics implemented as Excel sheets
## Author:  Joseph T. Foley <foley AT RU DOT IS>
## Start Date: 2022-09-12
## Input: Spreadsheet and updated deduction values "CON1 = 2"
## Output:  Updated Spreadsheet
## Ubuntu install
##   sudo apt install python3-openpyxl
## Doc: https://openpyxl.readthedocs.io/en/stable/tutorial.html 
import os
from pathlib import PurePath
import argparse
import logging
from datetime import datetime
from collections import OrderedDict
import re
from openpyxl import load_workbook

class ExcelRubric():
    """Handle all Excel operations."""
    def __init__(self, filepath, logger, lastrow=100, startval=0):
        self.logger = logger
        self.wb = load_workbook(filename=filepath)
    ## TODO:  figure out how to deal with CODE heading and number or just CODES

        ws = self.wb.active # TODO need to change to sheet 0 
        firstcol = []
        
        for row in range(1, lastrow):
            val = ws.cell(column=1, row=row).value
            if val is None:
                next
            else:
                firstcol.append(val)
        self.logger.debug(firstcol)
        # convert for orderdict for later
        codes = OrderedDict()
        for k in firstcol:
            codes[k] = startval
        self.codes = codes
    def dump_values(self,outfd):
        '''Given a file descriptor, iterate line by line output the database'''
        keyvals = sorted(self.codes.items())
        for (key, val) in keyvals:
            print(f"{key} {val}", file=outfd)
        
    def get_codes(self):
        return(self.codes)

def main():
    """Main program loop"""
    print("""Excel Rubric Manager/Abstraction by Joseph. T. Foley<foley AT ru DOT is>
    From https://github.com/foleyj2/teaching-tools""")
    parser = argparse.ArgumentParser(
        description="Load XLS file and manipulate.")
    parser.add_argument('filepath')
    parser.add_argument('--log', default="INFO",
        help='Console log level:  Number or DEBUG, INFO, WARNING, ERROR')
    parser.add_argument('--ext', default=".xlr",
                        help='Extension for output')

    args = parser.parse_args()
    ## Set up logging
    numeric_level = getattr(logging, args.log.upper(), None)
    if not isinstance(numeric_level, int):
        raise ValueError('Invalid log level: %s' % args.log)
    print(f"Log level:  {numeric_level}")
    logger = logging.getLogger("app")
    logger.setLevel(logging.INFO)
    # log everything to file
    ## TODO:  common logger for all modules
    logpath = 'ExcelRubric-{:%Y%m%d-%H%M%S}.log'.format(datetime.now())
    fh = logging.FileHandler(logpath)
    fh.setLevel(logging.DEBUG)
    # log to console
    ch = logging.StreamHandler()
    ch.setLevel(numeric_level)
    # create formatter and add to handlers
    consoleformatter = logging.Formatter('%(message)s')
    ch.setFormatter(consoleformatter)
    spamformatter = logging.Formatter('%(asctime)s %(name)s[%(levelname)s] %(message)s')
    fh.setFormatter(spamformatter)
    # add the handlers to logger
    logger.addHandler(ch)
    logger.addHandler(fh)

    logger.info("Creating ExcelRubric log file %s", logpath)
    # filename pre-processing for output
    inpath = PurePath(args.filepath)
    outpath = inpath.with_suffix(args.ext)
    print(f"File: {inpath} -> {outpath}")
    with open(outpath, "w") as outfd:
        ER = ExcelRubric(inpath, logger)
        AA.dump_values(outfd)
        print(ER.get_codes())
# STUB:  We know that codes are in column A
# STUB:  We know that we need to update values in column C

if __name__ == "__main__":
  main()
